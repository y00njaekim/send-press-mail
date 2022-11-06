from openpyxl import load_workbook

load_wb = load_workbook("/Users/yoonjae/Downloads/Book1.xlsx", data_only = True)
load_ws = load_wb['Sheet1']

for row in range(1,load_ws.max_row+1):
  cell_name = "B{}".format(row)
  name = load_ws[cell_name].value
  
  cell_email = "E{}".format(row)
  email = load_ws[cell_email].value